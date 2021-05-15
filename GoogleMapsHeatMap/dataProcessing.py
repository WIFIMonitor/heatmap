from __future__ import print_function
import time
import swagger_client
import requests
import json
import openpyxl
import math
import schedule
import time
from bullet import VerticalPrompt, Password, Input
from pathlib import Path
from swagger_client.rest import ApiException
from pprint import pprint
from requests.auth import HTTPBasicAuth
from influxdb import InfluxDBClient
from datetime import datetime, timedelta

# -------------------------------- Saving data from the xlsx file to a Dict ----------------------------------

# Function to read the data given from the xlsx file
def readXlsx(dir, fileName):
    # Setting the path to the xlsx file:
    xlsx_file = Path(dir, fileName)

    # Read the Excel File
    wb_obj = openpyxl.load_workbook(xlsx_file)

    # Read the Active Sheet from the Excel file
    sheet = wb_obj.active

    # Max rows
    # print(("Number of rows: %d") % (sheet.max_row))

    #accessPoints = {}
    i = 0
    aps={}
    fo = open("file2.txt","w")
    fo.write("[")
    for row in sheet.iter_rows(max_col=7, values_only=True):
        a = ""
        
        
        if (i != 0):
            if(row[3]!= None):
                #id = row[0]
                apData = {
                    'id' : row[0],
                    'latitude' : row[3],
                    'longitude' : row[4],
                }
                a = "{" + "id: '" + str(row[0]) + "', Latitude: " + str(row[3]) + ", Longitude: " + str(row[4]) + "},\n"
                fo.write(a)     
                #accessPoints[id] = apData
                aps[i] = apData


        i+=1
    fo.write("]")
    fo.close()
    print("file.txt created")
    return apData



xlsxData = readXlsx('.', 'PrimeCore.xlsx')

# ------------------------------------------- Functions ------------------------------------------------------
